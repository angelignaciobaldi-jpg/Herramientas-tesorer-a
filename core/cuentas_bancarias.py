"""Catálogo de cuentas bancarias de las empresas que dispersan.

Lee el Excel 'Cuentas bancarias/CUENTAS BANCARIAS .xlsx' (hoja con columnas
Alias/Empresa, Divisa, Banco, CLABE) y permite consultar la CLABE de cuenta
origen por (empresa, banco). Así el usuario elige empresa + banco y la cuenta
aparece sola, sin escribirla.

Si el Excel se actualiza, basta reabrir la aplicación para reflejar los cambios.
"""

from __future__ import annotations

import json
import os
import re

try:
    import openpyxl
except ImportError:  # openpyxl es opcional; sin él, el catálogo queda vacío.
    openpyxl = None

from . import rutas

# El Excel lo actualiza el usuario, así que va junto al .exe (no empaquetado).
RUTA_EXCEL = os.path.join(rutas.DATOS, "Cuentas bancarias", "CUENTAS BANCARIAS .xlsx")
# Copia en caché del catálogo. Permite seguir trabajando aunque el Excel esté
# abierto (Excel lo bloquea exclusivamente) usando la última lectura válida.
_RUTA_CACHE = os.path.join(rutas.DATOS, "_cuentas_cache.json")

# Banco elegido en la interfaz -> nombre del banco tal como aparece en el Excel.
BANCO_UI_A_EXCEL = {
    "Banregio": "BANREGIO",
    "Bancomer": "BBVA BANCOMER",
}


def _limpiar_clabe(valor) -> str:
    return re.sub(r"\D", "", str(valor or ""))


def _limpiar_cuenta(valor) -> str:
    # El número de cuenta es informativo; solo se quita el apóstrofo de Excel.
    return str(valor or "").replace("'", "").strip()


def _leer_excel(ruta: str) -> dict[str, dict[str, list[list[str]]]]:
    """Lee el Excel. Devuelve {} si no se puede (no existe, bloqueado, etc.).

    Columnas del Excel: A=Cuenta, B=Alias/Empresa, C=Divisa, D=Banco, E=CLABE.
    """
    catalogo: dict[str, dict[str, list[list[str]]]] = {}
    if openpyxl is None or not os.path.exists(ruta):
        return catalogo
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    except Exception:
        return catalogo  # p. ej. PermissionError si está abierto en Excel
    nombre_hoja = "RESUMEN DE CUENTAS" if "RESUMEN DE CUENTAS" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[nombre_hoja]
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila:
            continue
        cuenta = _limpiar_cuenta(fila[0]) if len(fila) > 0 else ""
        empresa = str(fila[1]).strip() if len(fila) > 1 and fila[1] else ""
        divisa = str(fila[2]).strip() if len(fila) > 2 and fila[2] else ""
        banco = str(fila[3]).strip() if len(fila) > 3 and fila[3] else ""
        clabe = _limpiar_clabe(fila[4]) if len(fila) > 4 else ""
        if not empresa or not banco or len(clabe) != 18:
            continue
        catalogo.setdefault(empresa, {}).setdefault(banco, []).append([clabe, divisa, cuenta])
    wb.close()
    return catalogo


def _cargar(ruta: str) -> dict[str, dict[str, list[list[str]]]]:
    """Carga el catálogo; si el Excel se puede leer, actualiza el caché; si no
    (p. ej. está abierto en Excel), usa la última lectura guardada en caché."""
    catalogo = _leer_excel(ruta)
    if catalogo:
        try:  # guarda la última versión buena
            with open(_RUTA_CACHE, "w", encoding="utf-8") as fh:
                json.dump(catalogo, fh, ensure_ascii=False)
        except Exception:
            pass
        return catalogo
    # No se pudo leer el Excel -> intentar el caché.
    if os.path.exists(_RUTA_CACHE):
        try:
            with open(_RUTA_CACHE, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except Exception:
            pass
    return {}


class CatalogoCuentas:
    """Acceso al catálogo de cuentas por empresa y banco."""

    def __init__(self, ruta: str = RUTA_EXCEL):
        self.datos = _cargar(ruta)

    def disponible(self) -> bool:
        return bool(self.datos)

    def empresas(self) -> list[str]:
        return sorted(self.datos.keys())

    def cuentas(self, empresa: str, banco_ui: str) -> list[tuple[str, str, str]]:
        """Cuentas (clabe, divisa, num_cuenta) de una empresa para el banco
        elegido en la UI. Las cuentas en PESOS/MXP se listan primero."""
        banco = BANCO_UI_A_EXCEL.get(banco_ui, banco_ui)
        cuentas = list(self.datos.get(empresa, {}).get(banco, []))
        cuentas.sort(key=lambda cd: 0 if cd[1].upper() in ("PESOS", "MXP") else 1)
        return cuentas
