"""Generación del reporte Excel de la dispersión de devoluciones."""

from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_AZUL = "1F4E78"
_GRIS = "D9D9D9"
_BORDE = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def generar(ruta: str, contexto: dict, registros: list[tuple]) -> None:
    """Crea el archivo Excel.

    Args:
        ruta: ruta destino .xlsx
        contexto: {empresa, banco, cuenta_origen, num_cuenta, fecha}
        registros: lista de (clabe, monto, beneficiario, concepto, dia)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Devoluciones"

    # --- Encabezado informativo ---
    ws["A1"] = "Reporte de dispersión de devoluciones"
    ws["A1"].font = Font(bold=True, size=14, color=_AZUL)
    info = [
        ("Empresa:", contexto.get("empresa", "")),
        ("Banco:", contexto.get("banco", "")),
        ("Cuenta origen (CLABE):", contexto.get("cuenta_origen", "")),
        ("Número de cuenta:", contexto.get("num_cuenta", "")),
        ("Fecha:", contexto.get("fecha", "")),
    ]
    fila = 3
    for etiqueta, valor in info:
        ws.cell(row=fila, column=1, value=etiqueta).font = Font(bold=True)
        ws.cell(row=fila, column=2, value=valor)
        fila += 1

    # --- Tabla de movimientos ---
    fila += 1
    encabezados = ["#", "CLABE", "Monto", "Beneficiario",
                   "Concepto / Referencia", "Fecha de devolución (día)"]
    for col, titulo in enumerate(encabezados, start=1):
        c = ws.cell(row=fila, column=col, value=titulo)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDE

    fila_inicio = fila + 1
    for i, (clabe, monto, beneficiario, concepto, dia) in enumerate(registros, start=1):
        valores = [i, clabe, float(monto or 0), beneficiario, concepto, dia]
        for col, valor in enumerate(valores, start=1):
            c = ws.cell(row=fila_inicio + i - 1, column=col, value=valor)
            c.border = _BORDE
            if col == 1 or col == 6:
                c.alignment = Alignment(horizontal="center")
            if col == 2:
                c.alignment = Alignment(horizontal="center")
            if col == 3:
                c.number_format = '#,##0.00'

    # --- Total de montos ---
    fila_total = fila_inicio + len(registros)
    ws.cell(row=fila_total, column=2, value="TOTAL").font = Font(bold=True)
    ct = ws.cell(row=fila_total, column=3,
                 value=f"=SUM(C{fila_inicio}:C{fila_total - 1})")
    ct.font = Font(bold=True)
    ct.number_format = '#,##0.00'
    ct.fill = PatternFill("solid", fgColor=_GRIS)

    # --- Anchos de columna ---
    anchos = {"A": 5, "B": 24, "C": 16, "D": 34, "E": 34, "F": 22}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    wb.save(ruta)
